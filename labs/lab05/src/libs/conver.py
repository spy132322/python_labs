# src/lab05/json_csv.py
from pathlib import Path
import json
import csv
from typing import List, Dict, Any


def _ensure_relative(path: Path) -> None:
    if path.is_absolute():
        raise ValueError("Путь должен быть относительным")


def json_to_csv(json_path: str | Path, csv_path: str | Path) -> None:
    """
    Преобразует JSON-файл в CSV.

    Поддерживается только JSON, содержащий список словарей: [{...}, {...}, ...].
    Порядок колонок: берётся порядок ключей из первого объекта списка (т.е. порядок
    сохранённый в первом dict). Если у последующих объектов отсутствует поле,
    в CSV подставляется пустая строка.

    Кодировка: UTF-8.

    Ошибки:
        - неверный тип файла (расширение) -> ValueError
        - путь абсолютный -> ValueError
        - отсутствующий файл -> FileNotFoundError (при открытии)
        - пустой JSON или неподдерживаемая структура -> ValueError
        - список содержит элементы не-словари -> ValueError
    """
    jp = Path(json_path)
    cp = Path(csv_path)

    _ensure_relative(jp)
    _ensure_relative(cp)

    if jp.suffix.lower() != ".json":
        raise ValueError("Неверный тип входного файла: ожидался .json")
    if cp.suffix.lower() != ".csv":
        raise ValueError("Неверный тип выходного файла: ожидался .csv")

    with jp.open("r", encoding="utf-8") as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError as e:
            raise ValueError(f"Невалидный JSON: {e}")

    if not isinstance(data, list) or len(data) == 0:
        raise ValueError("Пустой JSON или неподдерживаемая структура")

    if not all(isinstance(item, dict) for item in data):
        raise ValueError("Список должен содержать только словари")

    first = data[0]
    if not isinstance(first, dict) or len(first) == 0:
        headers: List[str] = list(first.keys())
    else:
        headers = list(first.keys())

    all_keys = set()
    for d in data:
        all_keys.update(d.keys())

    remaining = [k for k in sorted(all_keys) if k not in headers]
    headers.extend(remaining)

    if cp.parent and not cp.parent.exists():
        cp.parent.mkdir(parents=True, exist_ok=True)

    with cp.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        for row in data:
            out_row = {k: ("" if row.get(k) is None else str(row.get(k))) for k in headers}
            writer.writerow(out_row)


def csv_to_json(csv_path: str | Path, json_path: str | Path) -> None:
    """
    Преобразует CSV в JSON (список словарей).

    Первая строка CSV обязана быть заголовком. Все значения сохраняются как строки.
    Результат записывается с json.dump(..., ensure_ascii=False, indent=2).

    Кодировка: UTF-8.

    Ошибки:
        - неверный тип файла (расширение) -> ValueError
        - путь абсолютный -> ValueError
        - отсутствующий файл -> FileNotFoundError (при открытии)
        - CSV без заголовка или пустой -> ValueError
    """
    cp = Path(csv_path)
    jp = Path(json_path)

    _ensure_relative(cp)
    _ensure_relative(jp)

    if cp.suffix.lower() != ".csv":
        raise ValueError("Неверный тип входного файла: ожидался .csv")
    if jp.suffix.lower() != ".json":
        raise ValueError("Неверный тип выходного файла: ожидался .json")

    with cp.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError("CSV без заголовка или пустой")
        rows: List[Dict[str, str]] = []
        for raw in reader:
            row = {k: ("" if v is None else str(v)) for k, v in raw.items()}
            rows.append(row)

    if len(rows) == 0:
        raise ValueError("CSV без заголовка или пустой")

    if jp.parent and not jp.parent.exists():
        jp.parent.mkdir(parents=True, exist_ok=True)

    with jp.open("w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

from pathlib import Path
import csv
from typing import List

def csv_to_xlsx(csv_path: str | Path, xlsx_path: str | Path) -> None:
    """
    Конвертирует CSV в XLSX.

    Требования:
        - Использует openpyxl (если не установлен, бросает ImportError с подсказкой).
        - Первая строка CSV — заголовок.
        - Лист называется "Sheet1".
        - Колонки получают автоширину по длине текста (минимум 8).
        - Кодировка CSV: UTF-8.

    Ошибки:
        - неверный тип файла (расширение) -> ValueError
        - путь абсолютный -> ValueError
        - отсутствующий файл -> FileNotFoundError (при открытии)
        - CSV без заголовка или пустой -> ValueError
    """
    cp = Path(csv_path)
    xp = Path(xlsx_path)

    _ensure_relative(cp)
    _ensure_relative(xp)

    if cp.suffix.lower() != ".csv":
        raise ValueError("Неверный тип входного файла: ожидался .csv")
    if xp.suffix.lower() not in (".xlsx",):
        raise ValueError("Неверный тип выходного файла: ожидался .xlsx")
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError(
            "Для конвертации CSV в XLSX требуется пакет 'openpyxl'. "
            "Установите его: pip install openpyxl"
        ) from e

    with cp.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        try:
            rows = list(reader)
        except csv.Error as e:
            raise ValueError(f"Ошибка при чтении CSV: {e}")

    if not rows:
        raise ValueError("CSV без заголовка или пустой")

    header = rows[0]
    if not header or all(h == "" for h in header):
        raise ValueError("CSV без заголовка или пустой")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for r in rows:
        ws.append([("" if c is None else str(c)) for c in r])

    num_cols = max(len(r) for r in rows)
    col_max_lengths: List[int] = [0] * num_cols
    for r in rows:
        for i in range(num_cols):
            val = ""
            if i < len(r) and r[i] is not None:
                val = str(r[i])
            l = len(val)
            if l > col_max_lengths[i]:
                col_max_lengths[i] = l

    MIN_WIDTH = 8
    for i, maxlen in enumerate(col_max_lengths, start=1):
        width = max(MIN_WIDTH, maxlen)
        width = float(width) + 1.5
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    if xp.parent and not xp.parent.exists():
        xp.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(xp))
