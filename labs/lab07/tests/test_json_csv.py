from pathlib import Path
import json
import csv
import pytest
from libs.json_csv import json_to_csv, csv_to_json


def write_json(path: Path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def read_csv_rows(path: Path):
    with path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def test_csv_to_json_roundtrip(tmp_path: Path):
    src = tmp_path / "people.csv"
    dst = tmp_path / "people.json"
    header = ["name", "age"]
    rows = [["Alice", "22"], ["Bob", "25"]]
    with src.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)

    csv_to_json(str(src), str(dst))

    data = read_json(dst)
    assert isinstance(data, list)
    assert len(data) == 2
    assert set(data[0].keys()) == {"name", "age"}
    assert data[0]["name"] == "Alice"


def test_json_to_csv_empty_input_raises(tmp_path: Path):
    src = tmp_path / "empty.json"
    src.write_text("", encoding="utf-8")
    dst = tmp_path / "out.csv"

    with pytest.raises(ValueError):
        json_to_csv(str(src), str(dst))


def test_csv_to_json_raises_on_missing_header(tmp_path: Path):
    src = tmp_path / "no_header.csv"
    src.write_text("1,2,3\n4,5,6\n", encoding="utf-8")
    dst = tmp_path / "out.json"
    with pytest.raises(ValueError):
        csv_to_json(str(src), str(dst))


def test_csv_to_json_raises_on_empty_header_name(tmp_path: Path):
    src = tmp_path / "empty_header.csv"
    src.write_text(",age\nAlice,22\n", encoding="utf-8")
    dst = tmp_path / "out.json"
    with pytest.raises(ValueError):
        csv_to_json(str(src), str(dst))


def test_csv_to_json_raises_on_duplicate_header_names(tmp_path: Path):
    src = tmp_path / "dup_header.csv"
    src.write_text("name,age,name\nAlice,22,alice_alias\n", encoding="utf-8")
    dst = tmp_path / "out.json"
    with pytest.raises(ValueError):
        csv_to_json(str(src), str(dst))


def test_nonexistent_path_raises(tmp_path: Path):
    with pytest.raises(FileNotFoundError):
        json_to_csv(str(tmp_path / "no_such.json"), str(tmp_path / "out.csv"))

    with pytest.raises(FileNotFoundError):
        csv_to_json(str(tmp_path / "no_such.csv"), str(tmp_path / "out.json"))


def test_roundtrip_preserves_fieldset_order_independence(tmp_path: Path):
    src = tmp_path / "people2.json"
    dst_csv = tmp_path / "people2.csv"
    dst_json = tmp_path / "people2_out.json"
    data = [
        {"age": 30, "name": "Charlie"},
        {"name": "Dana", "age": 28},
    ]
    write_json(src, data)
    json_to_csv(str(src), str(dst_csv))
    csv_to_json(str(dst_csv), str(dst_json))

    out = read_json(dst_json)
    assert {"name", "age"} <= set(out[0].keys())
    assert len(out) == 2


def test_json_to_csv_roundtrip_minimal(tmp_path: Path):
    src = tmp_path / "people.json"
    dst = tmp_path / "people.csv"
    data = [{"name": "Alice", "age": 22}, {"name": "Bob", "age": 25}]
    src.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    json_to_csv(str(src), str(dst))
    with dst.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 2
    assert {"name", "age"} <= set(rows[0].keys())


def test_json_to_csv_invalid_extension(tmp_path: Path):
    src = tmp_path / "people.txt"
    dst = tmp_path / "out.csv"
    src.write_text("not json", encoding="utf-8")
    with pytest.raises(ValueError):
        json_to_csv(str(src), str(dst))


def test_json_to_csv_absolute_path_rejected(tmp_path: Path):
    src = tmp_path / "people.json"
    dst = tmp_path / "out.csv"
    src.write_text("[]", encoding="utf-8")
    with pytest.raises(ValueError):
        json_to_csv(str(src.resolve()), str(dst))


def test_json_to_csv_malformed_json(tmp_path: Path):
    src = tmp_path / "bad.json"
    dst = tmp_path / "out.csv"
    src.write_text("{ bad json", encoding="utf-8")
    with pytest.raises(ValueError):
        json_to_csv(str(src), str(dst))


def test_json_to_csv_empty_or_not_list(tmp_path: Path):
    src = tmp_path / "empty.json"
    dst = tmp_path / "out.csv"
    src.write_text("{}", encoding="utf-8")
    with pytest.raises(ValueError):
        json_to_csv(str(src), str(dst))

    src.write_text("[]", encoding="utf-8")
    with pytest.raises(ValueError):
        json_to_csv(str(src), str(dst))


def test_json_to_csv_non_dict_elements(tmp_path: Path):
    src = tmp_path / "mixed.json"
    dst = tmp_path / "out.csv"
    src.write_text(json.dumps([{"a": 1}, 2, "x"], ensure_ascii=False), encoding="utf-8")
    with pytest.raises(ValueError):
        json_to_csv(str(src), str(dst))


def test_csv_to_json_basic_roundtrip(tmp_path: Path):
    src = tmp_path / "people.csv"
    dst = tmp_path / "out.json"
    with src.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["name", "age"])
        w.writerow(["Alice", "22"])
    csv_to_json(str(src), str(dst))
    out = json.loads(dst.read_text(encoding="utf-8"))
    assert isinstance(out, list) and out[0]["name"] == "Alice"


def test_csv_to_json_missing_header_is_valid_by_default(tmp_path: Path):
    src = tmp_path / "no_header.csv"
    src.write_text("1,2,3\n4,5,6\n", encoding="utf-8")
    dst = tmp_path / "out.json"
    try:
        csv_to_json(str(src), str(dst))
    except ValueError:
        return
    out = json.loads(dst.read_text(encoding="utf-8"))
    assert out == [{"1": "4", "2": "5", "3": "6"}]


def test_csv_to_json_empty_header_name_raises(tmp_path: Path):
    src = tmp_path / "empty_header.csv"
    src.write_text(",age\nAlice,22\n", encoding="utf-8")
    dst = tmp_path / "out.json"
    with pytest.raises(ValueError):
        csv_to_json(str(src), str(dst))


def test_csv_to_json_duplicate_header_raises(tmp_path: Path):
    src = tmp_path / "dup.csv"
    src.write_text("name,age,name\nAlice,22,alias\n", encoding="utf-8")
    dst = tmp_path / "out.json"
    with pytest.raises(ValueError):
        csv_to_json(str(src), str(dst))


def test_csv_to_json_csv_without_rows_raises(tmp_path: Path):
    src = tmp_path / "head_only.csv"
    src.write_text("name,age\n", encoding="utf-8")
    dst = tmp_path / "out.json"
    with pytest.raises(ValueError):
        csv_to_json(str(src), str(dst))


def test_json_to_csv_first_element_empty_dict(tmp_path: Path):
    """Тест для случая, когда первый элемент списка - пустой словарь"""
    src = tmp_path / "empty_first.json"
    dst = tmp_path / "out.csv"
    data = [{}]
    write_json(src, data)

    with pytest.raises(ValueError, match="Пустой JSON или неподдерживаемая структура"):
        json_to_csv(str(src), str(dst))


def test_json_to_csv_none_values(tmp_path: Path):
    """Тест обработки None значений"""
    src = tmp_path / "with_none.json"
    dst = tmp_path / "out.csv"
    data = [{"name": "Alice", "age": None}, {"name": "Bob", "age": 25}]
    write_json(src, data)

    json_to_csv(str(src), str(dst))
    rows = read_csv_rows(dst)

    assert rows[0]["age"] == ""
    assert rows[1]["age"] == "25"


def test_csv_to_json_invalid_identifier_headers(tmp_path: Path):
    """Тест на невалидные идентификаторы в заголовках CSV"""
    src = tmp_path / "invalid_headers.csv"
    dst = tmp_path / "out.json"

    src.write_text("first name,age\nAlice,22\n", encoding="utf-8")

    with pytest.raises(ValueError, match="имена колонок должны быть идентификаторами"):
        csv_to_json(str(src), str(dst))


def test_csv_to_json_varying_columns(tmp_path: Path):
    """Тест CSV с разным количеством колонок в строках"""
    src = tmp_path / "varying.csv"
    dst = tmp_path / "out.json"

    content = """name,age,city
Alice,22,London
Bob,25
Charlie,30,Paris,extra
"""
    src.write_text(content, encoding="utf-8")

    csv_to_json(str(src), str(dst))
    data = read_json(dst)

    assert len(data) == 3
    assert data[0] == {"name": "Alice", "age": "22", "city": "London"}
    assert data[1] == {"name": "Bob", "age": "25", "city": ""}
    assert data[2] == {"name": "Charlie", "age": "30", "city": "Paris"}


def test_csv_to_xlsx_basic(tmp_path: Path):
    """Базовый тест конвертации CSV в XLSX"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        pytest.skip("openpyxl не установлен")

    src = tmp_path / "test.csv"
    dst = tmp_path / "test.xlsx"

    content = """name,age,city
Alice,22,London
Bob,25,New York
Charlie,30,Paris"""
    src.write_text(content, encoding="utf-8")

    from libs.json_csv import csv_to_xlsx

    csv_to_xlsx(str(src), str(dst))

    wb = load_workbook(dst)
    ws = wb.active

    assert ws.title == "Sheet1"
    assert ws["A1"].value == "name"
    assert ws["B1"].value == "age"
    assert ws["C1"].value == "city"
    assert ws["A2"].value == "Alice"
    assert ws["B3"].value == "25"


def test_csv_to_xlsx_empty_file(tmp_path: Path):
    """Тест пустого CSV файла"""
    src = tmp_path / "empty.csv"
    dst = tmp_path / "test.xlsx"
    src.write_text("", encoding="utf-8")

    from libs.json_csv import csv_to_xlsx

    with pytest.raises(ValueError, match="CSV без заголовка или пустой"):
        csv_to_xlsx(str(src), str(dst))


def test_csv_to_xlsx_no_header(tmp_path: Path):
    """Тест CSV без заголовка"""
    src = tmp_path / "no_header.csv"
    dst = tmp_path / "test.xlsx"
    src.write_text("Alice,22\nBob,25\n", encoding="utf-8")

    from libs.json_csv import csv_to_xlsx

    with pytest.raises(ValueError, match="CSV без заголовка или пустой"):
        csv_to_xlsx(str(src), str(dst))


def test_csv_to_xlsx_wrong_extension(tmp_path: Path):
    """Тест неверного расширения файлов"""
    src = tmp_path / "test.txt"
    dst = tmp_path / "test.xlsx"
    src.write_text("name,age\nAlice,22\n", encoding="utf-8")

    from libs.json_csv import csv_to_xlsx

    with pytest.raises(ValueError, match="Неверный тип входного файла"):
        csv_to_xlsx(str(src), str(dst))

    src = tmp_path / "test.csv"
    dst = tmp_path / "test.txt"
    src.write_text("name,age\nAlice,22\n", encoding="utf-8")

    with pytest.raises(ValueError, match="Неверный тип выходного файла"):
        csv_to_xlsx(str(src), str(dst))


def test_csv_to_xlsx_missing_file(tmp_path: Path):
    """Тест отсутствующего файла"""
    src = tmp_path / "missing.csv"
    dst = tmp_path / "test.xlsx"

    from libs.json_csv import csv_to_xlsx

    with pytest.raises(FileNotFoundError):
        csv_to_xlsx(str(src), str(dst))


def test_csv_to_xlsx_column_widths(tmp_path: Path):
    """Тест автоширины колонок"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        pytest.skip("openpyxl не установлен")

    src = tmp_path / "wide.csv"
    dst = tmp_path / "test.xlsx"

    content = """short,very_long_column_name,medium
a,this_is_a_very_long_value,test"""
    src.write_text(content, encoding="utf-8")

    from libs.json_csv import csv_to_xlsx

    csv_to_xlsx(str(src), str(dst))

    wb = load_workbook(dst)
    ws = wb.active

    assert ws.column_dimensions["A"].width >= 8
    assert ws.column_dimensions["B"].width > 20
